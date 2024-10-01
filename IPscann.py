"""
Network Scanner Script using ARP Table.

This script scans the network and logs details
about the devices on the network, including their IP address, MAC address,
hostname, and vendor (manufacturer). The information is saved in an Excel file with
columns for time, IP, MAC, hostname, and vendor.

Creator: Brandon Magaña Avalos
"""

import subprocess
import re
import socket
import requests
import time
from datetime import datetime
import openpyxl
from openpyxl.utils.cell import get_column_letter
import platform
import uuid
import os

# Set the working directory to the script's directory
script_dir = os.path.dirname(os.path.abspath(__file__))
os.chdir(script_dir)

def get_local_ip():
    """
    Get the local IP address of the current machine.

    Returns:
        str: The local IP address of the machine where the script is executed.
    """
    hostname = socket.gethostname()
    local_ip = socket.gethostbyname(hostname)
    return local_ip

def get_local_mac():
    """
    Get the local MAC address of the current machine.

    Returns:
        str: The local MAC address (not Wi-Fi MAC).
    """
    mac = uuid.getnode()
    mac_address = ':'.join(('%012X' % mac)[i:i+2] for i in range(0, 12, 2))
    return mac_address

def get_hostname(ip):
    """
    Resolve an IP address to its corresponding hostname.

    Args:
        ip (str): The IP address of the device.

    Returns:
        str: The hostname of the device if resolvable, otherwise 'Unknown'.
    """
    try:
        hostname = socket.gethostbyaddr(ip)[0]
    except socket.herror:
        hostname = None
    return hostname

def get_mac_vendor(mac):
    """
    Get the vendor/manufacturer name associated with a MAC address using an external API.

    Args:
        mac (str): The MAC address of the device.

    Returns:
        str: The vendor name, or 'Unknown' if the vendor is not available.
    """
    try:
        url = f'https://api.macvendors.com/{mac}'
        response = requests.get(url)
        if response.status_code == 200:
            return response.text
        else:
            return 'Unknown'
    except Exception as e:
        return 'Unknown'

def scan_network_with_arp():
    """
    Scan the network using the ARP command to retrieve IP and MAC addresses of devices.

    Returns:
        list: A list of dictionaries, each containing 'ip' and 'mac' keys for network devices.
    """
    command = 'arp -a'
    result = subprocess.run(command, capture_output=True, text=True, shell=True)
    devices = []
    local_ip_octet = get_local_ip().split('.')[0]  # Get the first octet of the local IP
    for line in result.stdout.split('\n'):
        if re.match(r'^\s*\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}', line):  # Match any IP address
            parts = line.split()
            ip = parts[0]
            mac = parts[1]
            devices.append({'ip': ip, 'mac': mac})
    return devices

def log_devices(devices, log_file):
    """
    Log the scanned network devices' information into an Excel file.

    Args:
        devices (list): A list of dictionaries containing 'ip' and 'mac' of network devices.
        log_file (str): The path to the Excel file where the information will be logged.
    """
    header = ["Time", "IP", "MAC", "Hostname", "Vendor"]
    try:
        # Load the existing workbook
        workbook = openpyxl.load_workbook(log_file)
        sheet = workbook.active
    except FileNotFoundError:
        # Create a new workbook if the file doesn't exist
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        # Set the header row
        sheet.append(header)
    
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    # Log the local machine first
    current_device_ip = get_local_ip()
    current_device_mac = get_local_mac()
    current_device_hostname = get_hostname(current_device_ip) or 'Unknown'
    current_device_vendor = get_mac_vendor(current_device_mac) or 'Unknown'
    current_device_row = [timestamp, current_device_ip, current_device_mac, current_device_hostname, current_device_vendor]
    
    # Append the local machine info
    sheet.append(current_device_row)

    # Get the first three octets of the local IP address
    local_ip_prefix = '.'.join(get_local_ip().split('.')[:3])
    
    for device in devices:
        # Skip devices with MAC address ff-ff-ff-ff-ff-ff
        if device['mac'].lower() == 'ff-ff-ff-ff-ff-ff':
            continue
        # Skip devices with IP address that does not start with the local IP prefix
        if not device['ip'].startswith(local_ip_prefix):
            continue

        hostname = get_hostname(device['ip']) or 'Unknown'
        vendor = get_mac_vendor(device['mac']) or 'Unknown'
        row = [timestamp, device['ip'], device['mac'], hostname, vendor]
        sheet.append(row)

    # Adjust column widths
    for col_idx, col in enumerate(sheet.columns, start=1):
        max_length = max(len(str(cell.value)) for cell in col)
        sheet.column_dimensions[get_column_letter(col_idx)].width = max_length + 2
    
    # Add filter to the columns
    sheet.auto_filter.ref = f'A1:E{sheet.max_row}'
    
    # Save the workbook, retrying if a PermissionError occurs
    while True:
        try:
            workbook.save(log_file)
            print(f"Information logged to {log_file}")
            break
        except PermissionError:
            print(f"Permission denied: {log_file} is open. Retrying in 5 seconds...")
            time.sleep(5)

def WindowsScann():
    """
    Perform a network scan and log the results on a Windows machine.

    The network is scanned using ARP, and the details of devices on the network are saved
    to an Excel file. The scan repeats every 60 seconds.
    """
    log_file = f"network-log-{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    
    while True:
        print(f"Scanning network on {socket.gethostname()} | IP address: {get_local_ip()}")
        devices = scan_network_with_arp()
        
        # Log devices, including the local machine
        log_devices(devices, log_file)
        
        print("Waiting 60 seconds before the next scan...")
        time.sleep(60)  # Wait for 60 seconds before the next scan

if __name__ == '__main__':
    """
    Main function to determine the OS and execute the network scan for Windows.
    """
    print("Running the Scan Script on: " + platform.system())
    if platform.system() == 'Linux':
        print('Linux scanning is not implemented.')
    elif platform.system() == 'Windows':
        WindowsScann()
    else:
        print('This script supports only Linux and Windows (10/11).')
